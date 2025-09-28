"""
Excel helper for LMS Explorer
Handles Excel export functionality
"""

import os
import sys
from typing import List, Dict, Any, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExcelHelper:
    """Helper class for Excel operations"""
    
    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel functionality. Install with: pip install openpyxl")
        
        self.workbook = None
        self.worksheet = None
        self.current_row = 1
    
    def create_workbook(self, filename: str = None):
        """
        Create a new Excel workbook
        Args:
            filename: Optional filename to save the workbook
        """
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "LMS Data"
        self.current_row = 1
        
        if filename:
            self.save_workbook(filename)
    
    def save_workbook(self, filename: str):
        """
        Save the workbook to a file
        Args:
            filename: The filename to save to
        """
        if not self.workbook:
            raise ValueError("No workbook created. Call create_workbook() first.")
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        self.workbook.save(filename)
    
    def add_headers(self, headers: List[str]):
        """
        Add header row to the worksheet
        Args:
            headers: List of header strings
        """
        if not self.worksheet:
            raise ValueError("No worksheet created. Call create_workbook() first.")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col, value=header)
            
            # Style headers
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        self.current_row += 1
    
    def add_row(self, data: List[Any]):
        """
        Add a data row to the worksheet
        Args:
            data: List of data values
        """
        if not self.worksheet:
            raise ValueError("No worksheet created. Call create_workbook() first.")
        
        for col, value in enumerate(data, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col, value=value)
        
        self.current_row += 1
    
    def add_data_table(self, headers: List[str], data: List[List[Any]]):
        """
        Add a complete data table with headers and rows
        Args:
            headers: List of header strings
            data: List of data rows
        """
        self.add_headers(headers)
        
        for row in data:
            self.add_row(row)
    
    def auto_fit_columns(self):
        """Auto-fit column widths"""
        if not self.worksheet:
            return
        
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_users_to_excel(self, users: List[Any], filename: str):
        """
        Export users data to Excel
        Args:
            users: List of user objects
            filename: Output filename
        """
        self.create_workbook()
        
        headers = ["ID", "First Name", "Last Name", "Email", "Full Name", "Roles"]
        self.add_headers(headers)
        
        for user in users:
            row_data = [
                getattr(user, 'id', ''),
                getattr(user, 'first_name', ''),
                getattr(user, 'last_name', ''),
                getattr(user, 'email', ''),
                getattr(user, 'full_name', ''),
                ', '.join(getattr(user, 'roles', []))
            ]
            self.add_row(row_data)
        
        self.auto_fit_columns()
        self.save_workbook(filename)
    
    def export_courses_to_excel(self, courses: List[Any], filename: str):
        """
        Export courses data to Excel
        Args:
            courses: List of course objects
            filename: Output filename
        """
        self.create_workbook()
        
        headers = ["ID", "Name", "Category", "Enrolled Users", "User Groups"]
        self.add_headers(headers)
        
        for course in courses:
            category_name = getattr(course.category, 'name', '') if getattr(course, 'category', None) else ''
            enrolled_count = len(getattr(course, 'enrolled_users', []))
            groups_count = len(getattr(course, 'user_groups', []))
            
            row_data = [
                getattr(course, 'id', ''),
                getattr(course, 'name', ''),
                category_name,
                enrolled_count,
                groups_count
            ]
            self.add_row(row_data)
        
        self.auto_fit_columns()
        self.save_workbook(filename)
    
    def export_categories_to_excel(self, categories: List[Any], filename: str):
        """
        Export categories data to Excel
        Args:
            categories: List of category objects
            filename: Output filename
        """
        self.create_workbook()
        
        headers = ["ID", "Name", "Courses Count"]
        self.add_headers(headers)
        
        for category in categories:
            courses_count = len(getattr(category, 'courses', []))
            
            row_data = [
                getattr(category, 'id', ''),
                getattr(category, 'name', ''),
                courses_count
            ]
            self.add_row(row_data)
        
        self.auto_fit_columns()
        self.save_workbook(filename)
    
    def show_and_close(self):
        """Show Excel and close the helper"""
        if not self.workbook:
            return
        
        # Auto-fit columns before showing
        self.auto_fit_columns()
        
        # Save to temporary file and open it
        temp_filename = os.path.join(os.path.expanduser('~'), 'Desktop', 'LMS_Export.xlsx')
        try:
            self.save_workbook(temp_filename)
            
            # Open the file with default application
            if sys.platform == "win32":
                os.startfile(temp_filename)
            elif sys.platform == "darwin":  # macOS
                os.system(f"open '{temp_filename}'")
            else:  # Linux
                os.system(f"xdg-open '{temp_filename}'")
        except Exception as e:
            print(f"Failed to open Excel file: {e}")
        
        # Close workbook
        self.workbook.close()
        self.workbook = None
        self.worksheet = None
        self.current_row = 1


class ExcelWorkSpace:
    """Excel workspace class for backward compatibility"""
    
    def __init__(self):
        self.excel_helper = ExcelHelper()
        self.filename = None
    
    def create(self, filename: str = None):
        """Create Excel workspace"""
        self.filename = filename
        self.excel_helper.create_workbook(filename)
    
    def add_headers(self, headers: List[str]):
        """Add headers to worksheet"""
        self.excel_helper.add_headers(headers)
    
    def add_row(self, data: List[Any]):
        """Add row to worksheet"""
        self.excel_helper.add_row(data)
    
    def show_and_finally(self):
        """Show Excel and cleanup"""
        if not self.filename:
            self.excel_helper.show_and_close()
        else:
            self.excel_helper.auto_fit_columns()
            self.excel_helper.save_workbook(self.filename)
            self.excel_helper.workbook.close()


# Backward compatibility functions
def CreateExcelWorkspace() -> ExcelWorkSpace:
    """Create Excel workspace (backward compatibility)"""
    return ExcelWorkSpace()


def ExportUsersToExcel(users: List[Any], filename: str):
    """Export users to Excel (backward compatibility)"""
    helper = ExcelHelper()
    helper.export_users_to_excel(users, filename)


def ExportCoursesToExcel(courses: List[Any], filename: str):
    """Export courses to Excel (backward compatibility)"""
    helper = ExcelHelper()
    helper.export_courses_to_excel(courses, filename)


def ExportCategoriesToExcel(categories: List[Any], filename: str):
    """Export categories to Excel (backward compatibility)"""
    helper = ExcelHelper()
    helper.export_categories_to_excel(categories, filename)
